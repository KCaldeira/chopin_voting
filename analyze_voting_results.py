import pandas as pd
import numpy as np
import itertools

# ============================================
# 1. Load data
# ============================================

def load_scores(csv_path: str):
    """
    Load a CSV of the form:
      Contestant,Judge1,Judge2,...,JudgeN
      Name1,25,24,...,23
      Name2,NA,23,...,24

    Returns:
      contestants: list of contestant names
      judges: list of judge column names
      scores: DataFrame of numeric scores (rows=contestants, cols=judges)
      df: original DataFrame including 'Contestant'
    """
    df = pd.read_csv(csv_path, na_values=["NA"])
    contestants = df["Contestant"].tolist()
    judges = [c for c in df.columns if c != "Contestant"]
    scores = df[judges].astype(float)
    return contestants, judges, scores, df


# ============================================
# 2. Pairwise / Condorcet (Copeland)
# ============================================

def compute_pairwise(scores: pd.DataFrame, contestants, judges):
    """
    Compute pairwise wins matrix, ties, and comparison counts.

    wins[a,b] = #judges preferring a over b
    ties[a,b] = #judges giving same score to a and b
    comps[a,b] = #judges that scored both a and b (non-NA)
    """
    n = len(contestants)
    wins = np.zeros((n, n), dtype=int)
    ties = np.zeros((n, n), dtype=int)
    comps = np.zeros((n, n), dtype=int)

    for j in judges:
        col = scores[j]
        for a_idx, b_idx in itertools.combinations(range(n), 2):
            sa = col.iloc[a_idx]
            sb = col.iloc[b_idx]
            if pd.isna(sa) or pd.isna(sb):
                continue
            comps[a_idx, b_idx] += 1
            comps[b_idx, a_idx] += 1
            if sa > sb:
                wins[a_idx, b_idx] += 1
            elif sb > sa:
                wins[b_idx, a_idx] += 1
            else:
                ties[a_idx, b_idx] += 1
                ties[b_idx, a_idx] += 1

    return wins, ties, comps


def copeland_scores(wins, comps, contestants):
    """
    Copeland score: +1 per opponent you beat, -1 per opponent that beats you.

    Returns:
      ranked: list of (contestant, copeland_score), sorted descending
      condorcet: list of Condorcet winners (usually 0 or 1 names)
    """
    n = len(contestants)
    copeland = np.zeros(n, dtype=int)

    for a in range(n):
        for b in range(n):
            if a == b:
                continue
            if comps[a, b] == 0:
                continue  # no comparison, ignore
            if wins[a, b] > wins[b, a]:
                copeland[a] += 1
            elif wins[b, a] > wins[a, b]:
                copeland[a] -= 1
            # ties give 0

    order = sorted(range(n), key=lambda i: copeland[i], reverse=True)
    ranked = [(contestants[i], int(copeland[i])) for i in order]

    # Condorcet winner: beats every other candidate
    condorcet = []
    for a in range(n):
        beats_all = True
        for b in range(n):
            if a == b:
                continue
            if comps[a, b] == 0:
                continue
            if wins[a, b] <= wins[b, a]:
                beats_all = False
                break
        if beats_all:
            condorcet.append(contestants[a])

    return ranked, condorcet


# ============================================
# 3. IRV (ranked-choice) with fractional top ties
# ============================================

def total_raw_score(scores: pd.DataFrame, df: pd.DataFrame, contestant: str):
    """Sum of all non-NA scores for a given contestant."""
    row = scores.loc[df["Contestant"] == contestant].iloc[0]
    return row.dropna().sum()


def irv_fractional_top(scores: pd.DataFrame, df: pd.DataFrame, contestants, judges):
    """
    IRV-style elimination:
      - In each round, each judge gives 1 vote total.
      - Among remaining contestants they scored, those with max score
        share the vote equally (fractionally).
      - Candidate with lowest total votes is eliminated each round.
      - Tie-breaker: lowest total raw score, then alphabetical.

    Returns:
      winner: name of the IRV winner
      elimination_order: list of dicts with keys:
          'round', 'eliminated', 'votes'
    """
    remaining = set(contestants)
    elimination_order = []
    round_num = 1

    while True:
        votes = {c: 0.0 for c in remaining}
        active_judges = 0

        for j in judges:
            col = scores[j]
            sub = col[df["Contestant"].isin(remaining)].dropna()
            if sub.empty:
                continue
            active_judges += 1
            max_score = sub.max()
            top_idx = sub[sub == max_score].index
            share = 1.0 / len(top_idx)
            for idx in top_idx:
                cand = df.loc[idx, "Contestant"]
                votes[cand] += share

        if active_judges == 0 or not remaining:
            return None, elimination_order

        total_votes = float(active_judges)

        # Check majority
        winner = None
        for c in remaining:
            if votes[c] > total_votes / 2.0:
                winner = c
                break

        if winner is not None:
            # Record all remaining contestants (except winner) as eliminated in this round
            other_remaining = remaining - {winner}
            # Sort by votes (descending), then by total raw score (descending), then alphabetical
            sorted_remaining = sorted(
                other_remaining,
                key=lambda c: (-votes[c], -total_raw_score(scores, df, c), c)
            )
            for c in sorted_remaining:
                elimination_order.append(
                    {"round": round_num, "eliminated": c, "votes": votes[c]}
                )
            return winner, elimination_order

        # Eliminate lowest vote-getter
        min_vote = min(votes[c] for c in remaining)
        lowest = [c for c in remaining if votes[c] == min_vote]

        if len(lowest) > 1:
            lowest = sorted(
                lowest,
                key=lambda c: (total_raw_score(scores, df, c), c)
            )
        elim = lowest[0]
        elimination_order.append(
            {"round": round_num, "eliminated": elim, "votes": votes[elim]}
        )
        remaining.remove(elim)

        if len(remaining) == 1:
            last = list(remaining)[0]
            return last, elimination_order

        round_num += 1


# ============================================
# 4. Bottom-elimination (reverse runoff)
# ============================================

def bottom_elimination_fractional(scores: pd.DataFrame, df: pd.DataFrame, contestants, judges):
    """
    Reverse elimination:
      - In each round, for each judge, find their lowest score among remaining.
      - Everyone tied at that lowest score gets equal share of a "bottom vote".
      - Candidate with highest total bottom-votes is eliminated.
      - Tie-breaker: lowest total raw score, then alphabetical.

    Returns:
      winner: name of survivor
      elimination_order: list of dicts with keys:
          'round', 'eliminated', 'bottom_score'
    """
    remaining = set(contestants)
    elimination_order = []
    round_num = 1

    while len(remaining) > 1:
        bottom_counts = {c: 0.0 for c in remaining}

        for j in judges:
            col = scores[j]
            sub = col[df["Contestant"].isin(remaining)].dropna()
            if sub.empty:
                continue
            min_score = sub.min()
            bottom_idx = sub[sub == min_score].index
            share = 1.0 / len(bottom_idx)
            for idx in bottom_idx:
                cand = df.loc[idx, "Contestant"]
                bottom_counts[cand] += share

        max_bottom = max(bottom_counts[c] for c in remaining)
        candidates_max_bottom = [c for c in remaining if bottom_counts[c] == max_bottom]

        if len(candidates_max_bottom) > 1:
            candidates_max_bottom = sorted(
                candidates_max_bottom,
                key=lambda c: (total_raw_score(scores, df, c), c)
            )
        elim = candidates_max_bottom[0]
        elimination_order.append(
            {"round": round_num, "eliminated": elim, "bottom_score": bottom_counts[elim]}
        )
        remaining.remove(elim)
        round_num += 1

    winner = list(remaining)[0]
    return winner, elimination_order


# ============================================
# 5. First Past The Post (top picks)
# ============================================

def first_past_the_post_fractional(scores: pd.DataFrame, df: pd.DataFrame, contestants, judges):
    """
    For each judge:
      - Find max score among contestants they scored.
      - All contestants with that max share 1 vote (fractionally).
    Returns dict: {contestant: fptp_share}
    """
    fptp = {c: 0.0 for c in contestants}

    for j in judges:
        col = scores[j]
        valid = col.dropna()
        if valid.empty:
            continue
        max_score = valid.max()
        top_idx = valid[valid == max_score].index
        share = 1.0 / len(top_idx)
        for idx in top_idx:
            cand = df.loc[idx, "Contestant"]
            fptp[cand] += share

    return fptp


def first_past_the_post_equal(scores: pd.DataFrame, df: pd.DataFrame, contestants, judges):
    """
    For each judge:
      - Find max score among contestants they scored.
      - All contestants with that max each get 1 point (no sharing).
    Returns dict: {contestant: fptp_count}
    """
    fptp = {c: 0.0 for c in contestants}

    for j in judges:
        col = scores[j]
        valid = col.dropna()
        if valid.empty:
            continue
        max_score = valid.max()
        top_idx = valid[valid == max_score].index
        for idx in top_idx:
            cand = df.loc[idx, "Contestant"]
            fptp[cand] += 1.0

    return fptp


# ============================================
# 6. Diagnostics: judge top-score summary
# ============================================

def judge_top_summary(scores: pd.DataFrame, judges):
    """
    Returns DataFrame with:
      Judge, MaxScore, NumContestantsAtMax, MinScore, MeanScore, MedianScore, StdDev
    """
    rows = []
    for j in judges:
        col = scores[j]
        valid = col.dropna()
        if valid.empty:
            continue
        max_score = valid.max()
        min_score = valid.min()
        mean_score = valid.mean()
        median_score = valid.median()
        std_score = valid.std()
        count_max = int((valid == max_score).sum())
        rows.append({
            "Judge": j,
            "MaxScore": float(max_score),
            "MinScore": float(min_score),
            "MeanScore": float(mean_score),
            "MedianScore": float(median_score),
            "StdDev": float(std_score),
            "NumContestantsAtMax": count_max
        })
    return pd.DataFrame(rows)


# ============================================
# 7. Diagnostics: per-judge First Past The Post contributions for one contestant
# ============================================

def fptp_contributions_for_candidate(scores: pd.DataFrame,
                                      df: pd.DataFrame,
                                      judges,
                                      candidate: str):
    """
    For a given candidate, show judge-by-judge:
      - judge
      - judge's max score
      - candidate's score (or NA)
      - number of contestants at that max
      - candidate's fractional FPTP share (ties split)
      - candidate's equal FPTP count (1 if at max, else 0)
    Returns a DataFrame.
    """
    row = df[df["Contestant"] == candidate]
    if row.empty:
        raise ValueError(f"Candidate '{candidate}' not found in data.")
    row = row.iloc[0]

    records = []
    for j in judges:
        col = scores[j]
        valid = col.dropna()
        if valid.empty:
            continue
        max_score = valid.max()
        top_idx = valid[valid == max_score].index
        num_at_max = len(top_idx)
        share_fractional = 1.0 / num_at_max
        cand_score = row[j]
        cand_at_max = (not pd.isna(cand_score)) and (cand_score == max_score)
        contrib_fractional = share_fractional if cand_at_max else 0.0
        contrib_equal = 1.0 if cand_at_max else 0.0

        records.append({
            "Judge": j,
            "JudgeMaxScore": float(max_score),
            "CandidateScore": float(cand_score) if not pd.isna(cand_score) else np.nan,
            "NumContestantsAtMax": num_at_max,
            "FractionalShare": contrib_fractional,
            "EqualShare": contrib_equal
        })

    return pd.DataFrame(records)


# ============================================
# 8. Workflow control function
# ============================================

def analyze_voting_file(csv_path: str):
    """
    Run complete voting analysis on a single CSV file.

    Args:
        csv_path: Path to the CSV file containing voting data

    Returns:
        dict containing all analysis results
    """
    print(f"\n{'=' * 80}")
    print(f"Analyzing: {csv_path}")
    print(f"{'=' * 80}")

    contestants, judges, scores, df = load_scores(csv_path)

    # --- Pairwise / Condorcet ---
    wins, ties, comps = compute_pairwise(scores, contestants, judges)
    copeland_ranking, condorcet = copeland_scores(wins, comps, contestants)
    print("\nCopeland ranking (pairwise):")
    for name, score in copeland_ranking:
        print(f"  {name}: {score}")
    print("Condorcet winner(s):", condorcet)

    # --- IRV (top-down) ---
    irv_winner, irv_elims = irv_fractional_top(scores, df, contestants, judges)
    print("\nIRV winner:", irv_winner)
    print("IRV elimination order:")
    for r in irv_elims:
        print(f"  Round {r['round']:2d}: eliminated {r['eliminated']} (votes={r['votes']:.3f})")

    # --- Bottom-elimination (reverse) ---
    bottom_winner, bottom_elims = bottom_elimination_fractional(scores, df, contestants, judges)
    print("\nBottom-elimination winner:", bottom_winner)
    print("Bottom-elimination order:")
    for r in bottom_elims:
        print(f"  Round {r['round']:2d}: eliminated {r['eliminated']} (bottom_score={r['bottom_score']:.3f})")

    # --- First Past The Post metrics ---
    fptp_frac = first_past_the_post_fractional(scores, df, contestants, judges)

    print("\nFirst Past The Post (fractional, ties share 1):")
    for name, val in sorted(fptp_frac.items(), key=lambda x: x[1], reverse=True):
        print(f"  {name}: {val:.3f}")

    # --- Diagnostics: judges' top-score summary ---
    print("\nJudge top-score summary:")
    jt = judge_top_summary(scores, judges)
    print(jt.to_string(index=False))

    # Return results for potential further processing
    return {
        'csv_path': csv_path,
        'contestants': contestants,
        'judges': judges,
        'copeland_ranking': copeland_ranking,
        'condorcet': condorcet,
        'irv_winner': irv_winner,
        'irv_eliminations': irv_elims,
        'bottom_winner': bottom_winner,
        'bottom_eliminations': bottom_elims,
        'first_past_the_post_fractional': fptp_frac,
        'judge_top_summary': jt
    }


# ============================================
# 9. Stage transition analysis
# ============================================

def analyze_stage_transitions(all_results: list):
    """
    Analyze which contestants advanced between stages and how well each
    voting method predicted advancement.

    Args:
        all_results: List of result dictionaries in stage order

    Returns:
        dict with advancement data for each stage
    """
    import os

    stage_transitions = {}

    # Map filenames to simpler stage names
    stage_order = []
    for result in all_results:
        filename = os.path.basename(result['csv_path'])
        stage_name = filename.replace('_scores_NA.csv', '').replace('Chopin_', '')
        stage_order.append((stage_name, result))

    # For each stage (except the last), determine who advanced
    for i in range(len(stage_order) - 1):
        current_stage_name, current_result = stage_order[i]
        next_stage_name, next_result = stage_order[i + 1]

        current_contestants = set(current_result['contestants'])
        next_contestants = set(next_result['contestants'])

        # Who advanced
        advanced = current_contestants & next_contestants
        num_advanced = len(advanced)

        # For each voting method, get top N and compare
        stage_transitions[current_stage_name] = {
            'next_stage': next_stage_name,
            'advanced_set': advanced,
            'num_advanced': num_advanced,
            'total_contestants': len(current_contestants),
            'voting_method_analysis': {}
        }

        # Copeland: top N by ranking
        copeland_top_n = set([name for name, score in current_result['copeland_ranking'][:num_advanced]])
        copeland_agreement = len(copeland_top_n & advanced)
        copeland_disagreement = num_advanced - copeland_agreement

        # IRV: top N by ranking (winner is rank 1, etc.)
        irv_order = [current_result['irv_winner']] + [e['eliminated'] for e in reversed(current_result['irv_eliminations'])]
        irv_top_n = set(irv_order[:num_advanced])
        irv_agreement = len(irv_top_n & advanced)
        irv_disagreement = num_advanced - irv_agreement

        # Bottom Elimination: top N by ranking
        bottom_order = [current_result['bottom_winner']] + [e['eliminated'] for e in reversed(current_result['bottom_eliminations'])]
        bottom_top_n = set(bottom_order[:num_advanced])
        bottom_agreement = len(bottom_top_n & advanced)
        bottom_disagreement = num_advanced - bottom_agreement

        # First Past The Post: top N by fractional score
        fptp_sorted = sorted(current_result['first_past_the_post_fractional'].items(), key=lambda x: x[1], reverse=True)
        fptp_top_n = set([name for name, score in fptp_sorted[:num_advanced]])
        fptp_agreement = len(fptp_top_n & advanced)
        fptp_disagreement = num_advanced - fptp_agreement

        # Condorcet: check if Condorcet winner (if exists) advanced
        if current_result['condorcet']:
            condorcet_winner = current_result['condorcet'][0]
            condorcet_disagreement = 0 if condorcet_winner in advanced else 1
        else:
            # No Condorcet winner, so no prediction to disagree with
            condorcet_disagreement = 'N/A'

        stage_transitions[current_stage_name]['voting_method_analysis'] = {
            'copeland': {
                'top_n': copeland_top_n,
                'agreement': copeland_agreement,
                'disagreement': copeland_disagreement
            },
            'condorcet': {
                'disagreement': condorcet_disagreement
            },
            'irv': {
                'top_n': irv_top_n,
                'agreement': irv_agreement,
                'disagreement': irv_disagreement
            },
            'bottom_elimination': {
                'top_n': bottom_top_n,
                'agreement': bottom_agreement,
                'disagreement': bottom_disagreement
            },
            'first_past_the_post': {
                'top_n': fptp_top_n,
                'agreement': fptp_agreement,
                'disagreement': fptp_disagreement
            }
        }

    return stage_transitions


# ============================================
# 10. Cumulative voting analysis
# ============================================

def create_cumulative_scores(csv_files: list):
    """
    Create cumulative score datasets where each stage includes judges from all previous stages.

    Args:
        csv_files: List of CSV file paths in chronological order

    Returns:
        List of tuples: (stage_name, cumulative_df) for each stage
    """
    import os

    cumulative_datasets = []
    all_stage_data = []

    # Load all stages first
    for csv_path in csv_files:
        df = pd.read_csv(csv_path, na_values=["NA"])
        filename = os.path.basename(csv_path)
        stage_name = filename.replace('_scores_NA.csv', '').replace('Chopin_', '')
        all_stage_data.append((stage_name, df, csv_path))

    # Create cumulative datasets
    for i, (current_stage_name, current_df, current_path) in enumerate(all_stage_data):
        # Start with current stage contestants
        current_contestants = set(current_df['Contestant'].tolist())

        # Build cumulative dataframe starting with Contestant column
        cumulative_df = pd.DataFrame({'Contestant': sorted(current_contestants)})

        # Add judges from all stages up to and including current stage
        for j in range(i + 1):
            prev_stage_name, prev_df, prev_path = all_stage_data[j]

            # Get judge columns from this previous stage
            prev_judges = [c for c in prev_df.columns if c != 'Contestant']

            # Prefix judge names with stage to avoid conflicts
            for judge in prev_judges:
                prefixed_judge = f"{prev_stage_name}_{judge}"

                # Create a mapping from contestant to score for this judge
                judge_scores = {}
                for idx, row in prev_df.iterrows():
                    contestant = row['Contestant']
                    if contestant in current_contestants:
                        judge_scores[contestant] = row[judge]

                # Add this judge's scores to cumulative df
                cumulative_df[prefixed_judge] = cumulative_df['Contestant'].map(judge_scores)

        cumulative_datasets.append((current_stage_name, cumulative_df, current_path))

    return cumulative_datasets


# ============================================
# 11. Excel report generation
# ============================================

def generate_excel_report(results: list, output_path: str, stage_transitions: dict = None):
    """
    Generate Excel report with multiple sheets from analysis results.
    Each stage gets one sheet with all tables side-by-side in different columns.

    Args:
        results: List of result dictionaries from analyze_voting_file()
        output_path: Path to output Excel file
        stage_transitions: Dict with stage transition analysis (optional)
    """
    import os
    from datetime import datetime

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Create summary sheet first
        summary_data = []
        for result in results:
            filename = os.path.basename(result['csv_path'])
            stage_name = filename.replace('_scores_NA.csv', '').replace('Chopin_', '')

            # Format Condorcet winner with tie information
            if result['condorcet']:
                condorcet_str = result['condorcet'][0]
            else:
                # Show tied top scorers with (tied) after each name
                top_score = result['copeland_ranking'][0][1]
                tied_names = [name for name, score in result['copeland_ranking'] if score == top_score]
                if len(tied_names) > 1:
                    tied_with_labels = [f"{name} (tied)" for name in tied_names]
                    condorcet_str = ', '.join(tied_with_labels)
                else:
                    condorcet_str = 'None'

            row_data = {
                'Stage': stage_name,
                'Copeland Winner': result['copeland_ranking'][0][0],
                'Condorcet Winner': condorcet_str,
                'IRV Winner': result['irv_winner'],
                'Bottom-Elimination Winner': result['bottom_winner'],
                'Num Contestants': len(result['contestants']),
                'Num Judges': len(result['judges'])
            }

            # Add transition analysis if available
            if stage_transitions and stage_name in stage_transitions:
                trans = stage_transitions[stage_name]
                row_data['Advanced to Next'] = trans['num_advanced']
                row_data['Copeland Advancement Disagreement'] = trans['voting_method_analysis']['copeland']['disagreement']
                row_data['Condorcet Advancement Disagreement'] = trans['voting_method_analysis']['condorcet']['disagreement']
                row_data['IRV Advancement Disagreement'] = trans['voting_method_analysis']['irv']['disagreement']
                row_data['Bottom-Elim Advancement Disagreement'] = trans['voting_method_analysis']['bottom_elimination']['disagreement']
                row_data['FPTP Advancement Disagreement'] = trans['voting_method_analysis']['first_past_the_post']['disagreement']
            else:
                # Final stage has no transition
                row_data['Advanced to Next'] = 'N/A'
                row_data['Copeland Advancement Disagreement'] = 'N/A'
                row_data['Condorcet Advancement Disagreement'] = 'N/A'
                row_data['IRV Advancement Disagreement'] = 'N/A'
                row_data['Bottom-Elim Advancement Disagreement'] = 'N/A'
                row_data['FPTP Advancement Disagreement'] = 'N/A'

            summary_data.append(row_data)

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Create one sheet per stage with all tables in different columns
        for result in results:
            filename = os.path.basename(result['csv_path'])
            stage_name = filename.replace('_scores_NA.csv', '').replace('Chopin_', '')
            # Excel sheet names limited to 31 chars
            sheet_name = stage_name[:31] if len(stage_name) > 31 else stage_name

            # Get transition data for this stage if available
            trans_data = stage_transitions.get(stage_name) if stage_transitions else None
            advanced_set = trans_data['advanced_set'] if trans_data else set()

            col_offset = 0  # Track column position

            # 1. Copeland Rankings with Condorcet Winner info
            copeland_data = []
            for rank, (contestant, score) in enumerate(result['copeland_ranking'], start=1):
                row = {'Rank': rank, 'Contestant': contestant, 'Copeland Score': score}
                if trans_data:
                    # Check if in top N according to copeland
                    copeland_top_n = trans_data['voting_method_analysis']['copeland']['top_n']
                    actual_advanced = contestant in advanced_set
                    should_advance = contestant in copeland_top_n
                    row['Advanced'] = 'Y' if actual_advanced else 'N'
                copeland_data.append(row)

            copeland_df = pd.DataFrame(copeland_data)

            # Format Condorcet winner info
            if result['condorcet']:
                condorcet_display = result['condorcet'][0]
            else:
                # Show tied top scorers
                top_score = result['copeland_ranking'][0][1]
                tied_names = [name for name, score in result['copeland_ranking'] if score == top_score]
                if len(tied_names) > 1:
                    tied_with_labels = [f"{name} (tied)" for name in tied_names]
                    condorcet_display = ', '.join(tied_with_labels)
                else:
                    condorcet_display = 'None'

            # Add header row (title only)
            header_values = ['COPELAND RANKING'] + [''] * (len(copeland_df.columns) - 1)
            header_df = pd.DataFrame([header_values], columns=copeland_df.columns)
            copeland_with_header = pd.concat([header_df, copeland_df], ignore_index=True)
            copeland_with_header.to_excel(writer, sheet_name=sheet_name,
                                         startrow=0, startcol=col_offset, index=False)
            col_offset += len(copeland_df.columns) + 1  # Move to next section with gap

            # Store Condorcet winner for separate column
            condorcet_col_data = condorcet_display

            # 2. IRV Results - Show full ordering (no Votes column)
            # Build full ranking: winner is 1st, last eliminated is 2nd, etc.
            irv_rankings = []
            irv_order = [result['irv_winner']] + [e['eliminated'] for e in reversed(result['irv_eliminations'])]

            irv_rankings.append({
                'Rank': 1,
                'Contestant': result['irv_winner'],
                'Round': 'Winner'
            })
            if trans_data:
                irv_rankings[0]['Advanced'] = 'Y' if result['irv_winner'] in advanced_set else 'N'

            # Reverse elimination order to show final ranking
            for idx, elim in enumerate(reversed(result['irv_eliminations'])):
                row = {
                    'Rank': idx + 2,
                    'Contestant': elim['eliminated'],
                    'Round': elim['round']
                }
                if trans_data:
                    row['Advanced'] = 'Y' if elim['eliminated'] in advanced_set else 'N'
                irv_rankings.append(row)

            irv_ranking_df = pd.DataFrame(irv_rankings)
            irv_header_values = ['IRV RESULTS'] + [''] * (len(irv_ranking_df.columns) - 1)
            irv_header = pd.DataFrame([irv_header_values], columns=irv_ranking_df.columns)
            irv_combined = pd.concat([irv_header, irv_ranking_df], ignore_index=True)
            irv_combined.to_excel(writer, sheet_name=sheet_name,
                                 startrow=0, startcol=col_offset, index=False)
            col_offset += len(irv_ranking_df.columns) + 1

            # 3. Bottom Elimination Results - Show full ordering (no Bottom Score column)
            # Build full ranking: winner is 1st, last eliminated is 2nd, etc.
            bottom_rankings = []
            bottom_order = [result['bottom_winner']] + [e['eliminated'] for e in reversed(result['bottom_eliminations'])]

            bottom_rankings.append({
                'Rank': 1,
                'Contestant': result['bottom_winner'],
                'Round': 'Winner'
            })
            if trans_data:
                bottom_rankings[0]['Advanced'] = 'Y' if result['bottom_winner'] in advanced_set else 'N'

            # Reverse elimination order to show final ranking
            for idx, elim in enumerate(reversed(result['bottom_eliminations'])):
                row = {
                    'Rank': idx + 2,
                    'Contestant': elim['eliminated'],
                    'Round': elim['round']
                }
                if trans_data:
                    row['Advanced'] = 'Y' if elim['eliminated'] in advanced_set else 'N'
                bottom_rankings.append(row)

            bottom_ranking_df = pd.DataFrame(bottom_rankings)
            bottom_header_values = ['BOTTOM ELIMINATION'] + [''] * (len(bottom_ranking_df.columns) - 1)
            bottom_header = pd.DataFrame([bottom_header_values], columns=bottom_ranking_df.columns)
            bottom_combined = pd.concat([bottom_header, bottom_ranking_df], ignore_index=True)
            bottom_combined.to_excel(writer, sheet_name=sheet_name,
                                    startrow=0, startcol=col_offset, index=False)
            col_offset += len(bottom_ranking_df.columns) + 1

            # 4. First Past The Post (no FPTP Score column, sorted by score)
            fptp_data = []
            # Sort contestants by FPTP score for ranking
            sorted_contestants = sorted(result['contestants'],
                                       key=lambda c: result['first_past_the_post_fractional'][c],
                                       reverse=True)

            for rank, contestant in enumerate(sorted_contestants, start=1):
                row = {
                    'Rank': rank,
                    'Contestant': contestant
                }
                if trans_data:
                    row['Advanced'] = 'Y' if contestant in advanced_set else 'N'
                fptp_data.append(row)

            fptp_df = pd.DataFrame(fptp_data)
            fptp_header_values = ['FIRST PAST THE POST'] + [''] * (len(fptp_df.columns) - 1)
            fptp_header = pd.DataFrame([fptp_header_values], columns=fptp_df.columns)
            fptp_with_header = pd.concat([fptp_header, fptp_df], ignore_index=True)
            fptp_with_header.to_excel(writer, sheet_name=sheet_name,
                                     startrow=0, startcol=col_offset, index=False)
            col_offset += len(fptp_df.columns) + 1

            # 5. Condorcet Winner (separate column)
            condorcet_data = pd.DataFrame([
                ['CONDORCET WINNER'],
                [condorcet_col_data]
            ], columns=['Condorcet Winner'])
            condorcet_data.to_excel(writer, sheet_name=sheet_name,
                                   startrow=0, startcol=col_offset, index=False, header=False)
            col_offset += 2  # Condorcet column + gap

            # 6. Judge Summary
            judge_summary = result['judge_top_summary'].copy()
            # Create header row matching the number of columns
            header_values = ['JUDGE SUMMARY'] + [''] * (len(judge_summary.columns) - 1)
            judge_header = pd.DataFrame([header_values], columns=judge_summary.columns)
            judge_with_header = pd.concat([judge_header, judge_summary], ignore_index=True)
            judge_with_header.to_excel(writer, sheet_name=sheet_name,
                                      startrow=0, startcol=col_offset, index=False)

        # Apply formatting after all sheets are written
        workbook = writer.book
        from openpyxl.styles import Alignment

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 1. Word wrap the first row
            for cell in worksheet[1]:
                cell.alignment = Alignment(wrap_text=True)

            # 2. Center Y/N values in Advanced columns
            # Find all columns that have "Advanced" in the header (row 1)
            advanced_cols = []
            for col_idx, cell in enumerate(worksheet[1], start=1):
                if cell.value == 'Advanced':
                    advanced_cols.append(col_idx)

            # Apply center alignment to all cells in Advanced columns
            for col_idx in advanced_cols:
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    cell = row[col_idx - 1]
                    if cell.value in ['Y', 'N']:
                        cell.alignment = Alignment(horizontal='center')

            # 3. Set column width to 20 for contestant and judge name columns
            # Check row 1 (pandas column headers) for Contestant or Judge
            for col in worksheet.columns:
                col_letter = col[0].column_letter
                # Check row 1 for column header
                header_value = worksheet[f'{col_letter}1'].value
                if header_value in ['Contestant', 'Judge']:
                    worksheet.column_dimensions[col_letter].width = 20

    print(f"\nExcel report saved to: {output_path}")


# ============================================
# 10. Example usage
# ============================================

if __name__ == "__main__":
    import glob
    import os
    from datetime import datetime

    # Process all CSV files in data/input directory
    input_dir = "./data/input"
    output_dir = "./data/output"

    # Sort files in chronological stage order (Stage1, Stage2, Stage3, FinalStage)
    def stage_sort_key(filepath):
        filename = os.path.basename(filepath)
        if 'Stage1' in filename:
            return 1
        elif 'Stage2' in filename:
            return 2
        elif 'Stage3' in filename:
            return 3
        elif 'FinalStage' in filename:
            return 4
        else:
            return 99

    csv_files = sorted(glob.glob(os.path.join(input_dir, "*.csv")), key=stage_sort_key)

    if not csv_files:
        print(f"No CSV files found in {input_dir}")
    else:
        results = []
        for csv_path in csv_files:
            result = analyze_voting_file(csv_path)
            results.append(result)

        # Summary across all files
        print(f"\n{'=' * 80}")
        print("SUMMARY ACROSS ALL STAGES")
        print(f"{'=' * 80}")
        for result in results:
            filename = os.path.basename(result['csv_path'])
            print(f"\n{filename}:")
            print(f"  Copeland winner: {result['copeland_ranking'][0][0]}")
            print(f"  Condorcet winner: {result['condorcet'] if result['condorcet'] else 'None'}")
            print(f"  IRV winner: {result['irv_winner']}")
            print(f"  Bottom-elimination winner: {result['bottom_winner']}")

        # Analyze stage transitions
        stage_transitions = analyze_stage_transitions(results)

        # Generate standard Excel report with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"chopin_voting_analysis_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        generate_excel_report(results, excel_path, stage_transitions)

        # ===== CUMULATIVE ANALYSIS =====
        print(f"\n{'=' * 80}")
        print("CUMULATIVE ANALYSIS - Including judges from previous stages")
        print(f"{'=' * 80}")

        # Create cumulative datasets
        cumulative_datasets = create_cumulative_scores(csv_files)

        # Run analysis on cumulative data
        cumulative_results = []
        for stage_name, cumulative_df, original_path in cumulative_datasets:
            print(f"\nAnalyzing cumulative data for {stage_name}...")
            print(f"  Contestants: {len(cumulative_df)}")
            print(f"  Total judges (cumulative): {len(cumulative_df.columns) - 1}")

            # Convert cumulative df to same format as load_scores returns
            contestants = cumulative_df['Contestant'].tolist()
            judges = [c for c in cumulative_df.columns if c != 'Contestant']
            scores = cumulative_df[judges].astype(float)

            # Run the same analysis workflow
            wins, ties, comps = compute_pairwise(scores, contestants, judges)
            copeland_ranking, condorcet = copeland_scores(wins, comps, contestants)

            irv_winner, irv_elims = irv_fractional_top(scores, cumulative_df, contestants, judges)
            bottom_winner, bottom_elims = bottom_elimination_fractional(scores, cumulative_df, contestants, judges)

            fptp_frac = first_past_the_post_fractional(scores, cumulative_df, contestants, judges)
            jt = judge_top_summary(scores, judges)

            # Store results
            cumulative_results.append({
                'csv_path': original_path,  # Use original path for stage naming
                'contestants': contestants,
                'judges': judges,
                'copeland_ranking': copeland_ranking,
                'condorcet': condorcet,
                'irv_winner': irv_winner,
                'irv_eliminations': irv_elims,
                'bottom_winner': bottom_winner,
                'bottom_eliminations': bottom_elims,
                'first_past_the_post_fractional': fptp_frac,
                'judge_top_summary': jt
            })

        # Analyze cumulative stage transitions
        cumulative_transitions = analyze_stage_transitions(cumulative_results)

        # Generate cumulative Excel report
        cumulative_excel_filename = f"chopin_cumulative_voting_analysis_{timestamp}.xlsx"
        cumulative_excel_path = os.path.join(output_dir, cumulative_excel_filename)

        generate_excel_report(cumulative_results, cumulative_excel_path, cumulative_transitions)

        print(f"\n{'=' * 80}")
        print("BOTH REPORTS COMPLETED")
        print(f"{'=' * 80}")
        print(f"Standard report: {excel_filename}")
        print(f"Cumulative report: {cumulative_excel_filename}")
