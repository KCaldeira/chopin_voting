#!/usr/bin/env python3
"""
Normalized Score Analysis for Chopin Competition

This script analyzes judge scores by normalizing them (z-scores) and computing:
1. Mean and standard deviation of normalized scores for each contestant
2. Correlation coefficients between judges and with the normalized means
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill


def load_scores(csv_path):
    """
    Load contestant scores from CSV file.

    Returns:
        contestants: list of contestant names
        judges: list of judge names
        scores: dict mapping (contestant, judge) to score (or None for NA)
        df: original DataFrame
    """
    df = pd.read_csv(csv_path)
    contestants = df.iloc[:, 0].tolist()
    judges = df.columns[1:].tolist()

    scores = {}
    for _, row in df.iterrows():
        contestant = row.iloc[0]
        for judge in judges:
            val = row[judge]
            if pd.notna(val):
                scores[(contestant, judge)] = float(val)
            else:
                scores[(contestant, judge)] = None

    return contestants, judges, scores, df


def normalize_judge_scores(scores, contestants, judges):
    """
    Normalize each judge's scores using z-score normalization.
    For each judge: z = (score - mean) / std_dev

    Returns:
        normalized_scores: dict mapping (contestant, judge) to normalized score
        judge_stats: dict mapping judge to (mean, std_dev)
    """
    normalized_scores = {}
    judge_stats = {}

    for judge in judges:
        # Collect all scores for this judge (excluding NA)
        judge_scores = []
        for contestant in contestants:
            score = scores.get((contestant, judge))
            if score is not None:
                judge_scores.append(score)

        if len(judge_scores) > 0:
            mean = np.mean(judge_scores)
            std_dev = np.std(judge_scores, ddof=1) if len(judge_scores) > 1 else 1.0

            # Avoid division by zero
            if std_dev == 0:
                std_dev = 1.0

            judge_stats[judge] = (mean, std_dev)

            # Normalize each score for this judge
            for contestant in contestants:
                score = scores.get((contestant, judge))
                if score is not None:
                    normalized_scores[(contestant, judge)] = (score - mean) / std_dev
                else:
                    normalized_scores[(contestant, judge)] = None
        else:
            judge_stats[judge] = (0.0, 1.0)

    return normalized_scores, judge_stats


def calculate_contestant_stats(normalized_scores, contestants, judges):
    """
    Calculate mean and std dev of normalized scores for each contestant.

    Returns:
        contestant_stats: dict mapping contestant to (mean, std_dev, count)
    """
    contestant_stats = {}

    for contestant in contestants:
        # Collect all normalized scores for this contestant (excluding NA)
        contestant_norm_scores = []
        for judge in judges:
            norm_score = normalized_scores.get((contestant, judge))
            if norm_score is not None:
                contestant_norm_scores.append(norm_score)

        if len(contestant_norm_scores) > 0:
            mean = np.mean(contestant_norm_scores)
            std_dev = np.std(contestant_norm_scores, ddof=1) if len(contestant_norm_scores) > 1 else 0.0
            count = len(contestant_norm_scores)
            contestant_stats[contestant] = (mean, std_dev, count)
        else:
            contestant_stats[contestant] = (0.0, 0.0, 0)

    return contestant_stats


def calculate_judge_correlations(scores, normalized_scores, contestants, judges, contestant_stats):
    """
    Calculate correlation coefficients between judges and with normalized means.

    Returns:
        correlation_matrix: DataFrame with correlation coefficients
    """
    # Create a DataFrame with all judge scores and normalized mean
    data = {}

    # Add each judge's scores
    for judge in judges:
        judge_scores = []
        for contestant in contestants:
            score = scores.get((contestant, judge))
            judge_scores.append(score if score is not None else np.nan)
        data[judge] = judge_scores

    # Add normalized mean column
    norm_means = []
    for contestant in contestants:
        mean, _, _ = contestant_stats.get(contestant, (np.nan, np.nan, 0))
        norm_means.append(mean)
    data['Normalized Mean'] = norm_means

    df = pd.DataFrame(data)

    # Calculate correlation matrix
    correlation_matrix = df.corr()

    return correlation_matrix


def create_excel_report(all_results, output_path):
    """
    Generate Excel report with normalized score analysis.

    all_results is a list of tuples: (stage_name, contestants, judges, scores,
                                       normalized_scores, contestant_stats, correlation_matrix)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets for contestant statistics
    for stage_name, contestants, judges, scores, normalized_scores, contestant_stats, correlation_matrix in all_results:
        # Create contestant stats sheet
        ws = wb.create_sheet(title=f"{stage_name}")

        # Header
        ws.append(['Contestant', 'Normalized Mean', 'Normalized Std Dev', 'Number of Judges'])

        # Sort contestants by normalized mean (high to low)
        sorted_contestants = sorted(contestants,
                                   key=lambda c: contestant_stats.get(c, (0, 0, 0))[0],
                                   reverse=True)

        for contestant in sorted_contestants:
            mean, std_dev, count = contestant_stats.get(contestant, (0.0, 0.0, 0))
            ws.append([contestant, round(mean, 4), round(std_dev, 4), count])

        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Create sheets for correlation matrices
    for stage_name, contestants, judges, scores, normalized_scores, contestant_stats, correlation_matrix in all_results:
        ws = wb.create_sheet(title=f"{stage_name}_Corr")

        # Write correlation matrix
        ws.append([''] + list(correlation_matrix.columns))
        for idx in correlation_matrix.index:
            row = [idx] + [round(val, 4) if pd.notna(val) else 'NA'
                          for val in correlation_matrix.loc[idx]]
            ws.append(row)

        # Format header row and column
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="CCCCCC",
                                                           end_color="CCCCCC",
                                                           fill_type="solid")

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def analyze_stage(csv_path, stage_name):
    """
    Run complete normalized score analysis for a single stage.

    Returns:
        Tuple of (stage_name, contestants, judges, scores, normalized_scores,
                  contestant_stats, correlation_matrix)
    """
    print(f"\n{'='*60}")
    print(f"Analyzing {stage_name}")
    print(f"{'='*60}\n")

    # Load data
    contestants, judges, scores, df = load_scores(csv_path)

    # Normalize scores
    normalized_scores, judge_stats = normalize_judge_scores(scores, contestants, judges)

    # Calculate contestant statistics
    contestant_stats = calculate_contestant_stats(normalized_scores, contestants, judges)

    # Print contestant stats (sorted by mean)
    print(f"Contestant Statistics (Normalized Scores):")
    print(f"{'Contestant':<30} {'Mean':>10} {'Std Dev':>10} {'N Judges':>10}")
    print("-" * 65)

    sorted_contestants = sorted(contestants,
                               key=lambda c: contestant_stats.get(c, (0, 0, 0))[0],
                               reverse=True)

    for contestant in sorted_contestants:
        mean, std_dev, count = contestant_stats.get(contestant, (0.0, 0.0, 0))
        print(f"{contestant:<30} {mean:10.4f} {std_dev:10.4f} {count:10}")

    # Calculate correlations
    correlation_matrix = calculate_judge_correlations(scores, normalized_scores,
                                                      contestants, judges, contestant_stats)

    print(f"\n{stage_name} - Judge Correlation Matrix:")
    print(correlation_matrix.round(4))

    return (stage_name, contestants, judges, scores, normalized_scores,
            contestant_stats, correlation_matrix)


if __name__ == "__main__":
    # Define stage files
    stage_files = [
        ("./data/input/Chopin_Stage1_scores_NA.csv", "Stage1"),
        ("./data/input/Chopin_Stage2_scores_NA.csv", "Stage2"),
        ("./data/input/Chopin_Stage3_scores_NA.csv", "Stage3"),
        ("./data/input/Chopin_FinalStage_scores_NA.csv", "FinalStage"),
    ]

    # Analyze all stages
    all_results = []
    for csv_path, stage_name in stage_files:
        if Path(csv_path).exists():
            result = analyze_stage(csv_path, stage_name)
            all_results.append(result)
        else:
            print(f"Warning: File not found: {csv_path}")

    # Generate Excel report
    if all_results:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"./data/output/normalized_analysis_{timestamp}.xlsx"

        # Ensure output directory exists
        Path("./data/output").mkdir(parents=True, exist_ok=True)

        create_excel_report(all_results, output_path)
        print(f"\n{'='*60}")
        print(f"Analysis complete!")
        print(f"{'='*60}")
